"""Structure-preserving text extraction.

The plan-verbatim version of this module flattens everything to a single string,
losing page/paragraph/section structure that downstream RAG (chunker + search +
citation) needs. This module instead emits a list of ``ExtractedBlock`` records
carrying the structural locators the chunker rolls into chunk metadata so that
final search results can cite "page 3, §Pricing" rather than "chunk 47".

The public surface is two functions:

- :func:`extract_blocks` — primary, returns ``list[ExtractedBlock]``. Use this.
- :func:`extract_text` — back-compat thin wrapper returning a flat string.
  Joins block texts with ``\\n\\n``. Kept for callers that don't yet consume
  structured blocks; new code should call :func:`extract_blocks`.

Per-format notes:
  - PDF: PyMuPDF ``page.get_text("dict")`` gives blocks/lines/spans with font
    sizes; heading detection uses span font size relative to page median.
  - DOCX: python-docx exposes ``paragraph.style.name``; ``Heading N`` styles
    become section_heading transitions.
  - XLSX: openpyxl rows-as-blocks; sheet name acts as section_heading.
  - TXT/CSV: double-newline-separated paragraphs; no headings.
  - MD: like TXT, but ``#``/``##`` lines update section_heading.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import fitz  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Public data shapes                                                          #
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ExtractedBlock:
    """One structurally-meaningful chunk of source text, with locators.

    ``paragraph_index`` is monotonic and document-wide (0-indexed); the chunker
    uses ranges of these to express a chunk's span. ``page`` is 1-indexed (PDF
    only; ``None`` for formats without natural pagination). ``section_heading``
    is the most recently seen heading at or before this block, propagated
    forward by the extractor so each block can be cited standalone.

    ``source_locator`` carries format-specific extras (e.g. ``{"sheet":
    "2024 Revenue", "row": 17}`` for XLSX, ``{"bbox": [...]}`` for PDF) that
    don't fit the common schema but may matter for richer UIs later.
    """

    text: str
    page: int | None
    paragraph_index: int
    section_heading: str | None
    source_locator: dict = field(default_factory=dict)


# --------------------------------------------------------------------------- #
# Public API                                                                  #
# --------------------------------------------------------------------------- #
def extract_blocks(file_path: str) -> list[ExtractedBlock]:
    """Route to the right per-format extractor by file extension."""
    ext = Path(file_path).suffix.lower()
    extractors: dict[str, Callable[[str], list[ExtractedBlock]]] = {
        ".pdf": _extract_pdf_blocks,
        ".docx": _extract_docx_blocks,
        ".xlsx": _extract_xlsx_blocks,
        ".txt": _extract_txt_blocks,
        ".md": _extract_md_blocks,
        ".csv": _extract_txt_blocks,
    }
    fn = extractors.get(ext)
    if fn is None:
        raise ValueError(f"Unsupported file type: {ext}")
    return fn(file_path)


def extract_text(file_path: str) -> str:
    """Back-compat flat-string variant. Joins block texts with double newlines."""
    blocks = extract_blocks(file_path)
    return "\n\n".join(b.text for b in blocks if b.text.strip())


# --------------------------------------------------------------------------- #
# PDF                                                                         #
# --------------------------------------------------------------------------- #
# A span's font size is compared against the page's median text-size; spans
# at least this multiplicative factor larger are treated as headings. The
# factor is conservative — most PDFs use 1.2-1.5x for headings; 1.25 catches
# the common cases without misclassifying merely-bold body text.
_PDF_HEADING_FONT_RATIO = 1.25
# Spans shorter than this many characters are too small to be a heading worth
# propagating (page numbers, watermarks, isolated punctuation).
_PDF_HEADING_MIN_CHARS = 3
# PyMuPDF block type codes: 0 = text, 1 = image.
_PDF_TEXT_BLOCK_TYPE = 0


def _extract_pdf_blocks(path: str) -> list[ExtractedBlock]:
    blocks: list[ExtractedBlock] = []
    paragraph_index = 0
    current_heading: str | None = None

    doc = fitz.open(path)
    try:
        for page_num, page in enumerate(doc, start=1):
            page_dict = page.get_text("dict")
            median_size = _pdf_page_median_size(page_dict)

            for raw_block in page_dict.get("blocks", []):
                if raw_block.get("type") != _PDF_TEXT_BLOCK_TYPE:
                    continue

                text, max_span_size = _pdf_block_text_and_max_size(raw_block)
                text = text.strip()
                if not text:
                    continue

                is_heading = (
                    median_size > 0
                    and max_span_size >= median_size * _PDF_HEADING_FONT_RATIO
                    and len(text) >= _PDF_HEADING_MIN_CHARS
                    and "\n" not in text  # genuine headings are one line
                )

                if is_heading:
                    current_heading = text
                    # Headings ARE emitted as blocks too — chunker can decide
                    # whether to include their text. Marking via locator lets
                    # downstream code distinguish "this chunk starts with the
                    # heading itself" from "this chunk inherits a heading".
                    blocks.append(ExtractedBlock(
                        text=text,
                        page=page_num,
                        paragraph_index=paragraph_index,
                        section_heading=current_heading,
                        source_locator={
                            "bbox": list(raw_block.get("bbox", ())),
                            "is_heading": True,
                        },
                    ))
                else:
                    blocks.append(ExtractedBlock(
                        text=text,
                        page=page_num,
                        paragraph_index=paragraph_index,
                        section_heading=current_heading,
                        source_locator={"bbox": list(raw_block.get("bbox", ()))},
                    ))
                paragraph_index += 1
    finally:
        doc.close()

    return blocks


def _pdf_page_median_size(page_dict: dict) -> float:
    sizes: list[float] = []
    for blk in page_dict.get("blocks", []):
        if blk.get("type") != _PDF_TEXT_BLOCK_TYPE:
            continue
        for line in blk.get("lines", []):
            for span in line.get("spans", []):
                size = span.get("size")
                if isinstance(size, (int, float)) and size > 0:
                    sizes.append(float(size))
    if not sizes:
        return 0.0
    sizes.sort()
    mid = len(sizes) // 2
    if len(sizes) % 2:
        return sizes[mid]
    return (sizes[mid - 1] + sizes[mid]) / 2.0


def _pdf_block_text_and_max_size(raw_block: dict) -> tuple[str, float]:
    """Concatenate spans within a block, capturing the largest font size seen."""
    line_texts: list[str] = []
    max_size = 0.0
    for line in raw_block.get("lines", []):
        span_texts: list[str] = []
        for span in line.get("spans", []):
            span_text = span.get("text", "")
            if span_text:
                span_texts.append(span_text)
            size = span.get("size", 0.0)
            if isinstance(size, (int, float)) and size > max_size:
                max_size = float(size)
        if span_texts:
            line_texts.append("".join(span_texts))
    return "\n".join(line_texts), max_size


# --------------------------------------------------------------------------- #
# DOCX                                                                        #
# --------------------------------------------------------------------------- #
def _extract_docx_blocks(path: str) -> list[ExtractedBlock]:
    blocks: list[ExtractedBlock] = []
    paragraph_index = 0
    current_heading: str | None = None

    doc = DocxDocument(path)
    for para in doc.paragraphs:
        text = (para.text or "").strip()
        if not text:
            continue

        style_name = (para.style.name if para.style else "") or ""
        is_heading = style_name.lower().startswith("heading") or style_name == "Title"

        if is_heading:
            current_heading = text
            blocks.append(ExtractedBlock(
                text=text,
                page=None,
                paragraph_index=paragraph_index,
                section_heading=current_heading,
                source_locator={"style": style_name, "is_heading": True},
            ))
        else:
            blocks.append(ExtractedBlock(
                text=text,
                page=None,
                paragraph_index=paragraph_index,
                section_heading=current_heading,
                source_locator={"style": style_name},
            ))
        paragraph_index += 1

    return blocks


# --------------------------------------------------------------------------- #
# XLSX                                                                        #
# --------------------------------------------------------------------------- #
def _extract_xlsx_blocks(path: str) -> list[ExtractedBlock]:
    """Rows-as-blocks. Sheet name acts as section_heading."""
    blocks: list[ExtractedBlock] = []
    paragraph_index = 0

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = ["" if c is None else str(c) for c in row]
                # Skip wholly-empty rows but preserve them for paragraph_index
                # continuity? No — chunker would just discard empty text anyway,
                # and skipping keeps paragraph_index dense.
                if not any(c.strip() for c in cells):
                    continue
                row_text = " | ".join(cells)
                blocks.append(ExtractedBlock(
                    text=row_text,
                    page=None,
                    paragraph_index=paragraph_index,
                    section_heading=sheet_name,
                    source_locator={"sheet": sheet_name, "row": row_idx},
                ))
                paragraph_index += 1
    finally:
        wb.close()

    return blocks


# --------------------------------------------------------------------------- #
# TXT / CSV / MD                                                              #
# --------------------------------------------------------------------------- #
def _extract_txt_blocks(path: str) -> list[ExtractedBlock]:
    """Double-newline-separated paragraphs. No headings."""
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        raw = fh.read()

    paragraphs = [p.strip() for p in raw.split("\n\n")]
    blocks: list[ExtractedBlock] = []
    paragraph_index = 0
    for para in paragraphs:
        if not para:
            continue
        blocks.append(ExtractedBlock(
            text=para,
            page=None,
            paragraph_index=paragraph_index,
            section_heading=None,
            source_locator={},
        ))
        paragraph_index += 1
    return blocks


def _extract_md_blocks(path: str) -> list[ExtractedBlock]:
    """Markdown: double-newline paragraphs + ATX headings (``#``/``##``/...)."""
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        raw = fh.read()

    paragraphs = [p.strip() for p in raw.split("\n\n")]
    blocks: list[ExtractedBlock] = []
    paragraph_index = 0
    current_heading: str | None = None

    for para in paragraphs:
        if not para:
            continue

        first_line = para.split("\n", 1)[0].strip()
        is_heading = first_line.startswith("#") and " " in first_line

        if is_heading:
            heading_text = first_line.lstrip("#").strip()
            current_heading = heading_text or current_heading
            blocks.append(ExtractedBlock(
                text=first_line,
                page=None,
                paragraph_index=paragraph_index,
                section_heading=current_heading,
                source_locator={"is_heading": True},
            ))
            paragraph_index += 1
            # If the paragraph has body content after the heading line, emit it too.
            if "\n" in para:
                body = para.split("\n", 1)[1].strip()
                if body:
                    blocks.append(ExtractedBlock(
                        text=body,
                        page=None,
                        paragraph_index=paragraph_index,
                        section_heading=current_heading,
                        source_locator={},
                    ))
                    paragraph_index += 1
        else:
            blocks.append(ExtractedBlock(
                text=para,
                page=None,
                paragraph_index=paragraph_index,
                section_heading=current_heading,
                source_locator={},
            ))
            paragraph_index += 1

    return blocks
